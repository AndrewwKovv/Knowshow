from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from datetime import datetime, timedelta
import os
import logging
import tempfile

logger = logging.getLogger(__name__)

async def export_products_to_excel(user_id, search_query, db_manager):
    """
    Экспортирует найденные товары в Excel и возвращает путь к файлу.
    Файл сохраняется в временной папке системы.
    """
    
    # We no longer keep a parsed-products pool. Export GlobalProduct entries
    # that match the query (case-insensitive substring search).
    q = (search_query or '').strip().lower()
    all_gps = await db_manager.get_global_products()
    parsed_products = [g for g in all_gps if q in (g.name or '').lower()]

    if not parsed_products:
        raise ValueError(f"Товары не найдены для экспорта по запросу '{search_query}'")
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Товары"
    
    # Заголовки
    headers = ["Название товара", "Продавец", "Цена (base_price, ₽)", "Оригинальная цена (₽)", "Ссылка", "Последняя фиксация"]
    ws.append(headers)
    
    # Форматирование заголовков
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
    
    # Заполняем данные
    # Export matching GlobalProduct records
    for product in parsed_products:
        ws.append([
            product.name,
            "",
            int(product.threshold_min) if product.threshold_min is not None else '',
            int(product.threshold_max) if product.threshold_max is not None else '',
            "",
            ''
        ])
    
    # Автоширина для столбцов
    column_widths = {
        'A': 30,  # Название
        'B': 20,  # Продавец
        'C': 18,  # Цена со скидкой
        'D': 18,  # Оригинальная цена
        'E': 40,  # Ссылка
        'F': 16   # Найдено
    }
    
    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width
    
    # Сохраняем файл во временной папке
    temp_dir = tempfile.gettempdir()
    filename = f"products_{search_query}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    filepath = os.path.join(temp_dir, filename)
    
    try:
        wb.save(filepath)
        return filepath
    
    except Exception as e:
        logger.error(f"Error saving Excel file: {e}")
        raise

def cleanup_export_file(filepath):
    """
    Удаляет файл экспорта после отправки пользователю
    """
    try:
        if os.path.exists(filepath):
            os.remove(filepath)
    except Exception as e:
        logger.error(f"Error cleaning up file {filepath}: {e}")


async def export_user_products_to_excel(user_id, db_manager):
    """
    Export GlobalProduct list for admin bulk editing. Only admins can export.
    Returns path to temporary file.
    """
    from openpyxl import Workbook
    import tempfile
    import os

    user = await db_manager.get_user(user_id)
    if not user or not user.is_admin:
        raise ValueError("Только админы могут экспортировать/редактировать глобальные товары")

    products = await db_manager.get_global_products()
    if not products:
        raise ValueError("Нет глобальных товаров для экспорта")

    wb = Workbook()
    ws = wb.active
    ws.title = "Глобальные товары"

    headers = ["Название", "Пороговая цена", "Слова исключения (через запятую)", "Ключевые слова (через запятую)"]
    ws.append(headers)

    for p in products:
        try:
            thr_min = p.threshold_min
            thr_max = p.threshold_max
        except Exception:
            thr_min = None
            thr_max = None

        if thr_min is not None and thr_max is not None:
            thr_cell = f"{int(thr_min)}-{int(thr_max)}"
        elif thr_max is not None:
            thr_cell = f"{int(thr_max)}"
        elif thr_min is not None:
            thr_cell = f"{int(thr_min)}"
        else:
            thr_cell = ""

        ws.append([
            p.name,
            thr_cell,
            (p.exclusions or ''),
            (p.keywords or '')
        ])

    for col in ['A', 'B', 'C', 'D']:
        ws.column_dimensions[col].width = 30

    temp_dir = tempfile.gettempdir()
    filename = f"global_products_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    filepath = os.path.join(temp_dir, filename)
    wb.save(filepath)
    return filepath

async def export_found_products_to_excel(product_name: str, found_products: list, scraper, site_base_discount: int = 11):
    """
    Экспортирует найденные товары в Excel с форматом:
    Название | Тип симки | Цена | Ссылка на ВБ
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from datetime import datetime
    import os
    import tempfile
    import logging

    logger = logging.getLogger(__name__)

    wb = Workbook()
    ws = wb.active
    ws.title = "Результаты"

    # Заголовки
    headers = ["Название товара", "Тип симки", "Цена (₽)", "Ссылка на ВБ"]
    ws.append(headers)

    # Форматирование заголовков
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    from openpyxl.styles import Alignment
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Заполняем данные
    for product in found_products:
        try:
            # Извлекаем информацию о товаре
            product_info = scraper.extract_product_info(product, user_discount=0)
            if not product_info:
                continue

            name = product_info.get('name', '')
            original_price = product_info.get('price', '')
            url = product_info.get('url', '')

            # Применяем глобальную скидку
            try:
                price_val = float(original_price or 0)
                discounted_price = int(round(price_val * (1 - float(site_base_discount) / 100.0)))
            except Exception:
                discounted_price = original_price

            # Определяем тип симки из метаданных или характеристик товара
            sim_type = _extract_sim_type(product)

            ws.append([name, sim_type, discounted_price, url])

        except Exception as e:
            logger.error(f"Error processing product for export: {e}")
            continue

    # Установка ширины столбцов
    ws.column_dimensions['A'].width = 40  # Название
    ws.column_dimensions['B'].width = 20  # Тип симки
    ws.column_dimensions['C'].width = 15  # Цена со скидкой
    ws.column_dimensions['D'].width = 50  # Ссылка

    # Выравнивание ячеек
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row:
            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        # Цена по центру
        row[2].alignment = Alignment(horizontal="center", vertical="center")

    # Сохраняем файл
    temp_dir = tempfile.gettempdir()
    safe_name = "".join([c for c in product_name if c.isalnum() or c in " _-"])[:50].strip()
    filename = f"export_{safe_name}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    filepath = os.path.join(temp_dir, filename)

    try:
        wb.save(filepath)
        logger.info(f"Excel file saved: {filepath}")
        return filepath
    except Exception as e:
        logger.error(f"Error saving Excel file: {e}")
        raise

def _extract_sim_type(product: dict) -> str:
    """
    Извлекает информацию о типе симки из характеристик товара
    Ищет в metadata.characteristics или в product fields
    """
    try:
        metadata = product.get('metadata') or product.get('meta') or {}
        characteristics = metadata.get('characteristics') or metadata.get('characteristicsList') or []

        if isinstance(characteristics, list):
            for char in characteristics:
                if not isinstance(char, dict):
                    continue
                char_name = (char.get('name') or '').lower()

                if 'sim' in char_name or 'сим' in char_name:
                    values = char.get('values') or char.get('value') or []
                    if isinstance(values, list) and len(values) > 0:
                        first_val = values[0]
                        if isinstance(first_val, dict):
                            return first_val.get('name') or first_val.get('value') or ''
                        elif isinstance(first_val, str):
                            return first_val
                    return ''

        product_name = (product.get('name') or '').lower()

        if 'esim' in product_name or 'e-sim' in product_name:
            if 'nano' in product_name or 'sim+' in product_name or '+sim' in product_name:
                return 'Nano-SIM + eSIM'
            return 'eSIM'
        elif 'nano' in product_name:
            return 'Nano-SIM'
        elif 'sim' in product_name:
            return 'SIM'

        return '-'

    except Exception as e:
        logging.getLogger(__name__).error(f"Error extracting SIM type: {e}")
        return '-'