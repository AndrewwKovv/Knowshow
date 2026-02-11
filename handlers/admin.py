from aiogram import Router, F
from aiogram.types import Message, Document, CallbackQuery, InlineKeyboardMarkup, InlineKeyboardButton
from aiogram.filters import Command
from aiogram.fsm.context import FSMContext
from aiogram.fsm.state import State, StatesGroup
from openpyxl import load_workbook
import json
import logging
import re
import asyncio

logger = logging.getLogger(__name__)
router = Router()

class AdminStates(StatesGroup):
    waiting_for_grant_user_id = State()
    waiting_for_revoke_user_id = State()
    waiting_for_site_discount = State()
    waiting_for_channel_id = State()
    waiting_for_price_update = State()

@router.message(Command("admin"))
async def admin_panel(message: Message, db_manager):
    """Панель администратора"""
    user = await db_manager.get_user(message.from_user.id)
    
    if not user or not user.is_admin:
        await message.answer("❌ У вас нет доступа к панели администратора")
        return
    # Show parser settings directly — no per-user management in this flow
    site_discount = await db_manager.get_setting('site_base_discount', '11')
    channel_id = await db_manager.get_setting('notification_channel_id', 'не установлен')

    text = f"""🔧 **Панель администратора**

⚙️ Настройки парсера

💰 Глобальная скидка: {site_discount}%
📢 ID канала уведомлений: {channel_id}
"""

    keyboard = InlineKeyboardMarkup(inline_keyboard=[
        [InlineKeyboardButton(text="💵 Обновить цены", callback_data="admin_update_prices")],
        [InlineKeyboardButton(text="🧹 Очистить товары", callback_data="admin_clear_tables")],
        [InlineKeyboardButton(text="🔄 Перезапустить парсер", callback_data="admin_restart_parser")],
        [InlineKeyboardButton(text="💰 Изменить скидку", callback_data="admin_set_site_discount")],
        [InlineKeyboardButton(text="📢 Установить ID канала", callback_data="admin_set_channel_id")],
        [InlineKeyboardButton(text="⬅️ Назад", callback_data="main_menu")],
    ])

    await message.answer(text, reply_markup=keyboard, parse_mode="Markdown")

# Removed per-user admin menu callback — admin panel shows parser settings directly via /admin

@router.callback_query(F.data == "admin_parser_settings")
async def admin_parser_settings(callback: CallbackQuery, db_manager):
    """Настройки парсера"""
    user = await db_manager.get_user(callback.from_user.id)
    if not user or not user.is_admin:
        await callback.answer("❌ Нет доступа", show_alert=True)
        return
    
    # Получаем текущие значения
    site_discount = await db_manager.get_setting('site_base_discount', '11')
    channel_id = await db_manager.get_setting('notification_channel_id', 'не установлен')
    
    text = f"""⚙️ **Настройки парсера**

💰 Глобальная скидка: {site_discount}%
📢 ID канала уведомлений: {channel_id}
"""
    
    keyboard = InlineKeyboardMarkup(inline_keyboard=[
        [InlineKeyboardButton(text="💰 Изменить скидку", callback_data="admin_set_site_discount")],
        [InlineKeyboardButton(text="📢 Установить ID канала", callback_data="admin_set_channel_id")],
        [InlineKeyboardButton(text="🧹 Очистить товары", callback_data="admin_clear_tables")],
        [InlineKeyboardButton(text="🔄 Перезапустить парсер", callback_data="admin_restart_parser")],
        [InlineKeyboardButton(text="⬅️ Назад", callback_data="main_menu")],
    ])
    
    await callback.message.edit_text(text, reply_markup=keyboard, parse_mode="Markdown")
    await callback.answer()

@router.callback_query(F.data == "admin_set_site_discount")
async def set_site_discount_prompt(callback: CallbackQuery, state: FSMContext, db_manager):
    """Запрос глобальной скидки"""
    current = await db_manager.get_setting('site_base_discount', '11')
    await callback.message.edit_text(
        f"💰 **Введите глобальную скидку для парсинга (0-100%)**\n\n"
        f"Текущее значение: {current}%",
        parse_mode="Markdown"
    )
    await state.set_state(AdminStates.waiting_for_site_discount)
    await callback.answer()

@router.message(AdminStates.waiting_for_site_discount)
async def handle_set_site_discount(message: Message, state: FSMContext, db_manager):
    """Установить глобальную скидку"""
    user = await db_manager.get_user(message.from_user.id)
    if not user or not user.is_admin:
        await message.answer("❌ Нет доступа")
        await state.clear()
        return
    
    try:
        # Accept formats like "11", "11%", " 11 % "
        txt = str(message.text or "").strip()
        txt = txt.rstrip('%').strip()
        # remove any non-digit characters
        digits = re.sub(r'[^0-9]', '', txt)
        if not digits:
            await message.answer("❌ Введите целое число, например: 11 или 11%")
            await state.clear()
            return
        val = int(digits)
        if not 0 <= val <= 100:
            await message.answer("❌ Значение должно быть от 0 до 100")
            return
        
        await db_manager.set_setting('site_base_discount', str(val))
        await message.answer(f"✅ Глобальная скидка парсинга установлена: **{val}%**", parse_mode="Markdown")
    except ValueError:
        await message.answer("❌ Введите целое число")
    finally:
        await state.clear()

@router.callback_query(F.data == "admin_set_channel_id")
async def set_channel_id_prompt(callback: CallbackQuery, state: FSMContext, db_manager):
    """Запрос ID канала для уведомлений"""
    current = await db_manager.get_setting('notification_channel_id', 'не установлен')
    text = f"""📢 **Введите ID канала для отправки найденных товаров**

Текущее значение: {current}

**Примеры:**
• Приватный канал: `-1001234567890`
• Публичный канал: `@channel_name` или `channel_name`

**Как получить ID приватного канала:**
1. Добавьте бота в канал как администратора
2. Отправьте в канал сообщение
3. Пошлите боту команду `/debugid`
4. Скопируйте ID из ответа
"""
    await callback.message.edit_text(text, parse_mode="Markdown")
    await state.set_state(AdminStates.waiting_for_channel_id)
    await callback.answer()

@router.message(AdminStates.waiting_for_channel_id)
async def handle_set_channel_id(message: Message, state: FSMContext, db_manager):
    """Установить ID канала"""
    user = await db_manager.get_user(message.from_user.id)
    if not user or not user.is_admin:
        await message.answer("❌ Нет доступа")
        await state.clear()
        return
    
    channel_id = message.text.strip()
    
    if not channel_id:
        await message.answer("❌ ID канала не может быть пустым")
        return
    
    try:
        await db_manager.set_setting('notification_channel_id', channel_id)
        await message.answer(f"✅ ID канала установлен: `{channel_id}`", parse_mode="Markdown")
    except Exception as e:
        await message.answer(f"❌ Ошибка: {e}")
    finally:
        await state.clear()

@router.callback_query(F.data == "admin_clear_tables")
async def admin_clear_tables(callback: CallbackQuery, db_manager):
    """Очистить товары"""
    user = await db_manager.get_user(callback.from_user.id)
    if not user or not user.is_admin:
        await callback.answer("❌ Нет доступа", show_alert=True)
        return

    try:
        del_gp = await db_manager.delete_all_global_products()
        await callback.message.edit_text(
            f"🧹 **Очищено:**\n"
            f"📦 Глобальные товары: {del_gp}",
            parse_mode="Markdown"
        )
        await callback.answer()
    except Exception as e:
        await callback.message.edit_text(f"❌ Ошибка очистки: {e}")
        await callback.answer()

@router.callback_query(F.data == "admin_users")
async def show_users(callback: CallbackQuery, db_manager):
    """Показать список пользователей"""
    user = await db_manager.get_user(callback.from_user.id)
    
    if not user or not user.is_admin:
        await callback.answer("❌ Нет доступа", show_alert=True)
        return
    
    async with db_manager.async_session() as session:
        from database.models import User
        from sqlalchemy import select
        result = await session.execute(select(User))
        users = result.scalars().all()
    
    text = "📋 **Список пользователей:**\n\n"
    for u in users:
        status = "✅" if u.has_access else "❌"
        admin_badge = "👑" if u.is_admin else ""
        text += f"`{u.telegram_id}` {status} {admin_badge} {u.username or 'нет имени'}\n"
    
    await callback.message.edit_text(text, parse_mode="Markdown")
    await callback.answer()
# Removed per-user management handlers (grant/revoke/list) — not needed in channel-driven flow

@router.message(F.document)
async def handle_bulk_upload(message: Message, state: FSMContext, db_manager):
    """Массовое добавление товаров из Excel (только админы)"""
    
    user = await db_manager.get_user(message.from_user.id)
    if not user or not user.is_admin:
        await message.answer("❌ Только админы могут загружать товары")
        return
    
    if not message.document.file_name.endswith(('xlsx', 'xls')):
        await message.answer("❌ Загрузите файл Excel (.xlsx или .xls)")
        return
    
    try:
        file = await message.bot.get_file(message.document.file_id)
        file_path = f"/tmp/{message.document.file_name}"
        await message.bot.download_file(file.file_path, file_path)
        
        wb = load_workbook(file_path)
        ws = wb.active
        
        products_added = 0
        errors = []
        
        # Удаляем все старые глобальные товары перед загрузкой
        await db_manager.delete_all_global_products()
        
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            try:
                name, threshold_cell, exclusions, keywords = row[:4]
                
                if not name or threshold_cell is None:
                    errors.append(f"Строка {row_idx}: отсутствует название или порог")
                    continue
                
                # Parse threshold
                try:
                    if isinstance(threshold_cell, (int, float)):
                        thr_min = float(threshold_cell)
                        thr_max = float(threshold_cell)
                    else:
                        s = str(threshold_cell).strip()
                        if '-' in s:
                            parts = [p.strip() for p in s.split('-', 1)]
                            thr_min = float(parts[0])
                            thr_max = float(parts[1])
                        else:
                            thr_min = 0.0
                            thr_max = float(s)
                except Exception as e:
                    errors.append(f"Строка {row_idx}: неверный формат порога")
                    continue
                
                # Parse exclusions & keywords with lowercase normalization
                exclusions_list = []
                if exclusions:
                    exclusions_list = [ex.strip().lower() for ex in str(exclusions).split(',') if ex and ex.strip()]
                
                keywords_list = []
                if keywords:
                    keywords_list = [kw.strip().lower() for kw in str(keywords).split(',') if kw and kw.strip()]
                
                # Add to global products
                await db_manager.add_global_product(
                    name=str(name).strip(),
                    threshold_min=thr_min,
                    threshold_max=thr_max,
                    keywords=json.dumps(keywords_list),
                    exclusions=json.dumps(exclusions_list)
                )
                products_added += 1
                
            except Exception as e:
                errors.append(f"Строка {row_idx}: {str(e)}")
        
        response = f"✅ **Добавлено товаров:** {products_added}\n"
        if errors:
            response += f"\n⚠️ **Ошибки:**\n" + "\n".join(errors[:5])
        
        await message.answer(response, parse_mode="Markdown")
        
        # Signal parser to run immediately
        try:
            import parser.signals as signals
            ev = getattr(signals, 'parse_event', None)
            if ev is not None:
                ev.set()
        except Exception:
            pass
        
    except Exception as e:
        await message.answer(f"❌ Ошибка при загрузке: {e}")

@router.callback_query(F.data == "admin_restart_parser")
async def admin_restart_parser(callback: CallbackQuery, db_manager):
    """Перезапустить парсер"""
    user = await db_manager.get_user(callback.from_user.id)
    if not user or not user.is_admin:
        await callback.answer("❌ Нет доступа", show_alert=True)
        return
    
    try:
        import parser.signals as signals
        restart_ev = getattr(signals, 'parser_restart_event', None)
        if restart_ev is not None:
            restart_ev.set()
            # Also trigger a parse event to make the parser reload products immediately
            try:
                ev = getattr(signals, 'parse_event', None)
                if ev is not None:
                    ev.set()
            except Exception:
                pass
            await callback.answer("✅ Парсер перезапускается и перечитывает товары...", show_alert=False)
        else:
            await callback.answer("❌ Не удалось найти сигнал перезагрузки", show_alert=True)
    except Exception as e:
        await callback.answer(f"❌ Ошибка: {e}", show_alert=True)

@router.callback_query(F.data == "admin_update_prices")
async def admin_update_prices_prompt(callback: CallbackQuery, state: FSMContext, db_manager):
    """Запрос на обновление цен"""
    user = await db_manager.get_user(callback.from_user.id)
    if not user or not user.is_admin:
        await callback.answer("❌ Нет доступа", show_alert=True)
        return
    
    text = """💵 **Обновление цен**

Отправьте список товаров в формате:

🇭🇰 Sim+eSim 17 Pro Max 512GB Blue — 134000₽
🇭🇰 Sim+eSim 17 Pro Max 512GB Orange — 128000₽

**Формат:**
• [ФЛАГ] [ТИП_SIM] [МОДЕЛЬ] [ПАМЯТЬ] [ЦВЕТ] — [ЦЕНА]₽

**Типы SIM:**
• Sim+eSim → nano-SIM+Esim (вычтем 8.5% от цены)
• eSim → Esim (вычтем 8.5% от цены)

**Бот:**
1. Сопоставит товары с БД (игнорируя флаги)
2. Вычтет 8.5% от цены
3. Установит конечную цену как потолок (threshold_max)
4. Установит начальную цену = конечная - 18000 (threshold_min)
"""
    await callback.message.edit_text(text, parse_mode="Markdown")
    await state.set_state(AdminStates.waiting_for_price_update)
    await callback.answer()

@router.message(AdminStates.waiting_for_price_update)
async def handle_price_update(message: Message, state: FSMContext, db_manager):
    """Обработка обновления цен"""
    user = await db_manager.get_user(message.from_user.id)
    if not user or not user.is_admin:
        await message.answer("❌ Нет доступа")
        await state.clear()
        return
    
    text_content = message.text.strip()
    
    if not text_content:
        await message.answer("❌ Сообщение пусто")
        await state.clear()
        return
    
    try:
        # Парсим сообщение
        entries = _parse_price_entries(text_content)
        
        if not entries:
            await message.answer("❌ Не удалось распарсить ни одного товара. Проверьте формат.")
            await state.clear()
            return
        
        # Получаем глобальные товары
        global_products = await db_manager.get_global_products()
        
        updated_count = 0
        not_found = []
        
        for entry in entries:
            # Ищем совпадение в БД
            matched_product = _find_matching_product(entry, global_products)
            
            if not matched_product:
                not_found.append(entry['original'])
                continue
            
            # Вычисляем цены: вычитаем 8.5% от объявленной цены
            try:
                final_price = int(round(float(entry['price']) * (1.0 - 0.085)))
            except Exception:
                final_price = int(entry['price'])

            min_price = final_price - 18000
            
            # Обновляем товар в БД
            try:
                await db_manager.add_global_product(
                    name=matched_product.name,
                    threshold_min=float(min_price),
                    threshold_max=float(final_price),
                    keywords=matched_product.keywords,
                    exclusions=matched_product.exclusions
                )
                updated_count += 1
                logger.info(f"Updated price for '{matched_product.name}': {min_price}-{final_price}")
            except Exception as e:
                logger.error(f"Error updating product {matched_product.name}: {e}")
        
        response = f"✅ **Обновление цен завершено**\n\n"
        response += f"📊 Обновлено товаров: **{updated_count}**\n"
        
        if not_found:
            response += f"\n Не найдено в БД ({len(not_found)}):\n"
            for item in not_found[:5]:
                response += f"• {item}\n"
            if len(not_found) > 5:
                response += f"• ... и ещё {len(not_found) - 5}\n"
        
        await message.answer(response, parse_mode="Markdown")
        
        # Перезапускаем парсер
        try:
            import parser.signals as signals
            ev = getattr(signals, 'parse_event', None)
            if ev is not None:
                ev.set()
                logger.info("Parser signalled after price update")
        except Exception:
            pass
    
    except Exception as e:
        logger.error(f"Error in price update: {e}", exc_info=True)
        await message.answer(f"❌ Ошибка: {e}")
    
    finally:
        await state.clear()

def _parse_price_entries(text: str) -> list:
    """Парсит список товаров из сообщения пользователя
    
    Формат:
    🇭🇰 Sim+eSim 17 Pro Max 512GB Blue — 134000₽
    """
    entries = []
    lines = text.strip().split('\n')
    
    for line in lines:
        line = line.strip()
        if not line or '—' not in line:
            continue
        
        try:
            # Разделяем по '—'
            parts = line.split('—')
            if len(parts) != 2:
                continue
            
            product_part = parts[0].strip()
            price_part = parts[1].strip()
            
            # Удаляем флаги стран (емодзи с кодами стран)
            product_part = re.sub(r'[\U0001F1E6-\U0001F1FF]+\s*', '', product_part).strip()
            
            # Извлекаем цену (ищем числа перед ₽)
            price_match = re.search(r'(\d+)\s*₽?', price_part)
            if not price_match:
                continue
            
            price = int(price_match.group(1))
            
            # Нормализуем типы SIM для сравнения
            sim_type = _normalize_sim_type(product_part)

            # If this is an "Air" model, force Esim: Air models don't ship as nano-SIM+Esim
            try:
                comps_tmp = _extract_components(product_part)
                model_tmp = comps_tmp.get('model')
                if model_tmp and 'air' in str(model_tmp).lower():
                    sim_type = 'Esim'
            except Exception:
                # keep parsed sim_type if anything goes wrong
                pass
            
            entries.append({
                'original': line,
                'product_text': product_part,
                'sim_type': sim_type,
                'price': price
            })
        
        except Exception as e:
            logger.warning(f"Failed to parse line '{line}': {e}")
            continue
    
    return entries

def _normalize_sim_type(text: str) -> str:
    """Нормализует тип SIM для сравнения
    
    Sim+eSim → nano-SIM+Esim
    eSim → Esim
    """
    text_lower = text.lower()
    
    # Заменяем Sim+eSim на nano-SIM+Esim
    if 'sim+esim' in text_lower or 'sim + esim' in text_lower:
        text = text.replace('Sim+eSim', 'nano-SIM+Esim').replace('sim+esim', 'nano-SIM+Esim')
        text = text.replace('Sim + eSim', 'nano-SIM+Esim').replace('sim + esim', 'nano-SIM+Esim')
    
    # Убедимся что eSim остаётся как Esim
    text = re.sub(r'e[Ss]im', 'Esim', text, flags=re.IGNORECASE)
    
    return text.strip()

def _extract_components(product_name: str) -> dict:
    """Извлекает компоненты названия товара
    
    Возвращает словарь с компонентами:
    - model: '17 Pro' (модель)
    - storage: '256GB' (память)
    - color: 'Blue' (цвет)
    - sim_type: 'Esim' или 'nano-SIM+Esim' (тип симкарты)
    
    Правило: если в названии нет 'esim', то по умолчанию это 'nano-SIM+Esim'
    """
    text = product_name.lower().strip()
    original = product_name.strip()
    
    # Удаляем iPhone если есть
    text = re.sub(r'\biphone\b', '', text, flags=re.IGNORECASE).strip()
    
    # ВАЖНО: Извлекаем тип SIM ДО нормализации (чтобы найти 'sim+esim')
    # По умолчанию — nano-SIM+Esim, если нет явного 'esim'
    sim_type = 'nano-SIM+Esim'  # Значение по умолчанию
    
    if 'sim+esim' in text or 'sim + esim' in text:
        sim_type = 'nano-SIM+Esim'
        text = re.sub(r'\s*sim\s*\+\s*esim\s*', ' ', text, flags=re.IGNORECASE).strip()
    elif 'esim' in text:
        sim_type = 'Esim'
        text = re.sub(r'\s*esim\s*', ' ', text, flags=re.IGNORECASE).strip()
    
    # Нормализуем пробелы
    text = re.sub(r'\s+', ' ', text).strip()
    
    # Извлекаем память (256GB, 512GB, 1TB, 2TB и т.д.)
    storage_match = re.search(r'(\d+(?:tb|gb))\b', text, re.IGNORECASE)
    storage = storage_match.group(1).upper() if storage_match else None
    if storage:
        text = re.sub(r'\s*' + re.escape(storage_match.group(0)) + r'\s*', ' ', text, flags=re.IGNORECASE).strip()
    
    # Нормализуем пробелы
    text = re.sub(r'\s+', ' ', text).strip()
    
    # Извлекаем модель (17 Pro, 17 Air, 17 Pro Max и т.д.)
    model_match = re.search(r'\b(\d+\s+(?:pro\s+max|pro|air|plus)?)\b', text, re.IGNORECASE)
    model = model_match.group(1).title() if model_match else None
    if model:
        text = re.sub(r'\b' + re.escape(model_match.group(0)) + r'\b', ' ', text, flags=re.IGNORECASE).strip()
    
    # Нормализуем пробелы
    text = re.sub(r'\s+', ' ', text).strip()
    
    # Оставшийся текст — это цвет
    color = text.strip().title() if text.strip() else None

    # Special-case: map 'White' -> 'Silver' for iPhone 17 Pro / 17 Pro Max
    try:
        if color and color.lower() == 'white' and model:
            model_lower = model.lower()
            if model_lower in ('17 pro', '17 pro max'):
                color = 'Silver'
    except Exception:
        # If anything goes wrong, keep original color
        pass
    
    return {
        'model': model,
        'storage': storage,
        'color': color,
        'sim_type': sim_type,
        'original': original
    }

def _components_match(entry_components: dict, product_components: dict) -> bool:
    """Проверяет совпадение всех компонентов товара
    
    Все компоненты должны совпадать:
    - Модель
    - Память
    - Цвет
    - Тип SIM
    """
    components_to_check = ['model', 'storage', 'color', 'sim_type']
    
    for component in components_to_check:
        entry_val = entry_components.get(component)
        product_val = product_components.get(component)
        
        if not entry_val or not product_val:
            if entry_val != product_val:
                return False
        
        if entry_val and product_val:
            if entry_val.lower() != product_val.lower():
                return False
    
    return True

def _find_matching_product(entry: dict, global_products: list):
    """Ищет товар в БД по сопоставлению компонентов
    
    Извлекает компоненты (модель, память, цвет, тип SIM) из сообщения
    пользователя и каждого товара в БД, затем сравнивает их.
    
    Игнорирует: iPhone, флаги стран
    """
    # Извлекаем компоненты из сообщения пользователя
    entry_components = _extract_components(entry['product_text'])
    
    for product in global_products:
        # Извлекаем компоненты из товара в БД
        product_components = _extract_components(product.name)
        
        # Проверяем совпадение всех компонентов
        if _components_match(entry_components, product_components):
            return product
    
    return None